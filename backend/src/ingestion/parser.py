from __future__ import annotations

import csv
import io
import json
import re
import warnings
from pathlib import Path
from typing import Literal

from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook

from src.models.document import RawDocument
from src.observability.logging import get_logger

log = get_logger("parser")


def parse_document(doc: RawDocument) -> str:
    file_type = doc.ref.file_type.lower()
    source_path = doc.ref.source_path

    try:
        if file_type in {".md", ".txt", ".rst", ""}:
            return _parse_text(doc.content)
        elif file_type == ".json":
            return _parse_json(doc.content)
        elif file_type == ".html" or file_type == ".htm":
            return _parse_html(doc.content)
        elif file_type == ".docx":
            return _parse_docx(doc.content)
        elif file_type == ".pdf":
            return _parse_pdf(doc.content)
        elif file_type == ".csv":
            return _parse_csv(doc.content)
        elif file_type in {".xlsx", ".xls"}:
            return _parse_excel(doc.content)
        elif file_type in {".py", ".js", ".ts", ".yaml", ".yml"}:
            return _parse_text(doc.content)
        else:
            log.warning("unknown_file_type", file_type=file_type, path=source_path)
            return _parse_text(doc.content)
    except Exception as e:
        log.error("parse_failed", path=source_path, error=str(e))
        return _parse_text(doc.content) if isinstance(doc.content, str) else ""


def _parse_text(content: bytes | str) -> str:
    if isinstance(content, bytes):
        return content.decode("utf-8", errors="replace")
    return content


def _parse_json(content: bytes | str) -> str:
    text = _parse_text(content)
    try:
        data = json.loads(text)
        parts = []

        if isinstance(data, dict):
            title = data.get("title", "")
            if title:
                parts.append(f"Title: {title}")

            body = data.get("body", data.get("description", data.get("content", "")))
            if body:
                parts.append(body)

            labels = data.get("labels", [])
            if labels:
                parts.append(f"Labels: {', '.join(str(l) for l in labels)}")

            assignees = data.get("assignees", [])
            if assignees:
                names = [a.get("name", str(a)) if isinstance(a, dict) else str(a) for a in assignees]
                parts.append(f"Assignees: {', '.join(names)}")

            state = data.get("state", "")
            if state:
                parts.append(f"State: {state}")

            return "\n\n".join(parts) if parts else text

        return text
    except json.JSONDecodeError:
        return text


def _parse_html(content: bytes | str) -> str:
    text = _parse_text(content)
    soup = BeautifulSoup(text, "html.parser")

    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()

    return soup.get_text(separator="\n", strip=True)


def _parse_docx(content: bytes | str) -> str:
    if isinstance(content, str):
        return content

    doc = DocxDocument(io.BytesIO(content))
    paragraphs = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            if para.style and para.style.name and para.style.name.startswith("Heading"):
                level = para.style.name.replace("Heading ", "").replace("Heading", "1")
                try:
                    hashes = "#" * int(level)
                except ValueError:
                    hashes = "#"
                paragraphs.append(f"{hashes} {text}")
            else:
                paragraphs.append(text)
    return "\n\n".join(paragraphs)


def _parse_pdf(content: bytes | str) -> str:
    if isinstance(content, str):
        return content

    try:
        pdf_module = _load_pdf_module()
        doc = pdf_module.open(stream=content, filetype="pdf")
        pages = []
        for page_num, page in enumerate(doc, 1):
            text = page.get_text().strip()
            if text:
                pages.append(f"[Page {page_num}]\n{text}")
        doc.close()
        return "\n\n".join(pages)
    except ImportError:
        log.warning("pdf_parser_unavailable")
        return ""


def _load_pdf_module():
    # PyMuPDF currently emits Python 3.12 deprecation warnings from its generated
    # SWIG bindings at import time. Keep the suppression scoped to that import.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"builtin type SwigPyPacked has no __module__ attribute",
            category=DeprecationWarning,
        )
        warnings.filterwarnings(
            "ignore",
            message=r"builtin type SwigPyObject has no __module__ attribute",
            category=DeprecationWarning,
        )
        warnings.filterwarnings(
            "ignore",
            message=r"builtin type swigvarlink has no __module__ attribute",
            category=DeprecationWarning,
        )
        import pymupdf

    return pymupdf


def _parse_csv(content: bytes | str) -> str:
    text = _parse_text(content)
    lines = text.strip().split("\n")
    if not lines:
        return ""

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return text

    headers = rows[0]
    output_parts = [f"Headers: {' | '.join(headers)}"]

    for row in rows[1:]:
        row_parts = []
        for header, value in zip(headers, row):
            if value.strip():
                row_parts.append(f"{header}: {value.strip()}")
        if row_parts:
            output_parts.append(" | ".join(row_parts))

    return "\n".join(output_parts)


def _parse_excel(content: bytes | str) -> str:
    if isinstance(content, str):
        return content

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets_output = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h) if h is not None else "" for h in rows[0]]
            sheet_parts = [f"Sheet: {sheet_name}", f"Headers: {' | '.join(headers)}"]

            for row in rows[1:]:
                row_parts = []
                for header, value in zip(headers, row):
                    if value is not None and str(value).strip():
                        row_parts.append(f"{header}: {value}")
                if row_parts:
                    sheet_parts.append(" | ".join(row_parts))

            sheets_output.append("\n".join(sheet_parts))

        wb.close()
        return "\n\n".join(sheets_output)
    except ImportError:
        log.warning("excel_parser_unavailable")
        return ""


def detect_doc_type(ref_path: str, content: str) -> str:
    path_lower = ref_path.lower()

    if "wiki" in path_lower:
        return "wiki"
    if "readme" in path_lower:
        return "readme"
    if "runbook" in path_lower or "playbook" in path_lower:
        return "runbook"
    if "incident" in path_lower:
        return "incident_report"
    if "meeting" in path_lower or "notes" in path_lower:
        return "meeting_notes"
    if "service" in path_lower and ("catalog" in path_lower or "matrix" in path_lower):
        return "service_catalog"
    if "architecture" in path_lower or "design" in path_lower:
        return "architecture_doc"
    if "issue" in path_lower:
        return "issue"
    if "merge_request" in path_lower or "mr-" in path_lower:
        return "merge_request"
    if ref_path.endswith((".xlsx", ".xls", ".csv")):
        return "spreadsheet"

    content_lower = content[:500].lower()
    if "incident" in content_lower:
        return "incident_report"
    if "runbook" in content_lower:
        return "runbook"
    if "meeting" in content_lower and "minutes" in content_lower:
        return "meeting_notes"

    return "unknown"
